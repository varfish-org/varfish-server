"""Tests for the ``file_export`` module."""

from datetime import timedelta
import gzip
import io
import tempfile
from unittest.mock import patch

from bgjobs.models import BackgroundJob
import django
from django.conf import settings
from django.utils import timezone
import openpyxl
from projectroles.models import Project
from requests_mock import Mocker
from test_plus.test import TestCase
from timeline.models import ProjectEvent

from clinvar.tests.factories import ClinvarFactory
from cohorts.tests.factories import TestCohortBase
from extra_annos.tests.factories import ExtraAnnoFactory, ExtraAnnoFieldFactory
from geneinfo.tests.factories import GnomadConstraintsFactory, RefseqToEnsemblFactory
from variants.tests.factories import (
    CaseWithVariantSetFactory,
    ExportProjectCasesFileBgJobFactory,
    ResubmitFormDataFactory,
    SmallVariantFactory,
)

from .. import file_export
from ..models import Case, CaseAwareProject, ExportFileBgJob, ExportProjectCasesFileBgJob


class MehariMockerMixin:
    def _set_mehari_mocker(self, mock_):
        for small_var in self.small_vars:
            mock_.get(
                (
                    "https://mehari.com/seqvars/csq?genome_release={genome_release}"
                    "&chromosome={chromosome}&position={position}&reference={reference}"
                    "&alternative={alternative}"
                ).format(
                    genome_release=small_var.release,
                    chromosome=small_var.chromosome,
                    position=small_var.start,
                    reference=small_var.reference,
                    alternative=small_var.alternative,
                ),
                status_code=200,
                json={
                    "version": {"tx_db": "0.25.1", "mehari": "0.25.4"},
                    "query": {
                        "genome_release": small_var.release,
                        "chromosome": small_var.chromosome,
                        "position": small_var.start,
                        "reference": small_var.reference,
                        "alternative": small_var.alternative,
                    },
                    "result": [
                        {
                            "feature_id": "NM_058167.2",
                            "consequences": ["three_prime_utr_exon_variant"],
                            "hgvs_p": "p.(=)",
                            "hgvs_t": "c.*60G>A",
                        },
                        {
                            "feature_id": "NM_194315.1",
                            "consequences": ["three_prime_utr_exon_variant"],
                            "hgvs_p": "p.(=)",
                            "hgvs_t": "c.*60G>A",
                        },
                    ],
                },
            )


class ExportTestBase(TestCase):
    """Base class for testing exports.

    Sets up the database fixtures for project, case, and small variants.
    """

    def setUp(self):
        self.superuser = self.make_user("superuser")
        self.case, self.variant_set, _ = CaseWithVariantSetFactory.get("small")
        self.small_vars = SmallVariantFactory.create_batch(3, variant_set=self.variant_set)
        self.small_vars.sort(key=lambda x: x.chromosome_no)
        self.extra_anno = [
            ExtraAnnoFactory(
                release=self.small_vars[0].release,
                chromosome=self.small_vars[0].chromosome,
                start=self.small_vars[0].start,
                end=self.small_vars[0].end,
                bin=self.small_vars[0].bin,
                reference=self.small_vars[0].reference,
                alternative=self.small_vars[0].alternative,
                anno_data=[9.89],
            ),
            ExtraAnnoFactory(
                release=self.small_vars[2].release,
                chromosome=self.small_vars[2].chromosome,
                start=self.small_vars[2].start,
                end=self.small_vars[2].end,
                bin=self.small_vars[2].bin,
                reference=self.small_vars[2].reference,
                alternative=self.small_vars[2].alternative,
                anno_data=[10.78],
            ),
        ]
        self.extra_anno_field = ExtraAnnoFieldFactory()
        self.bg_job = BackgroundJob.objects.create(
            name="job name",
            project=Project.objects.first(),
            job_type="variants.export_file_bg_job",
            user=self.superuser,
        )
        self.export_job = ExportFileBgJob.objects.create(
            project=self.bg_job.project,
            bg_job=self.bg_job,
            case=self.case,
            query_args={"export_flags": True, "export_comments": True},
            file_type="xlsx",
        )

        GnomadConstraintsFactory.reset_sequence()

        # Create two entries for first variant that is in clinvar (second variant in total)
        for small_var in self.small_vars:
            ClinvarFactory(
                release=small_var.release,
                chromosome=small_var.chromosome,
                start=small_var.start,
                end=small_var.end,
                bin=small_var.bin,
                reference=small_var.reference,
                alternative=small_var.alternative,
                summary_clinvar_review_status_label="criteria provided, single committer",
                summary_clinvar_pathogenicity_label="pathogenic",
                summary_clinvar_pathogenicity=["pathogenic"],
            )
            GnomadConstraintsFactory(ensembl_gene_id=small_var.ensembl_gene_id)
            RefseqToEnsemblFactory(
                entrez_id=small_var.refseq_gene_id,
                ensembl_gene_id=small_var.ensembl_gene_id,
                ensembl_transcript_id=small_var.ensembl_transcript_id,
            )


class CaseExporterTest(MehariMockerMixin, ExportTestBase):
    def setUp(self):
        super().setUp()
        # Here, the query arguments actually matter
        self.export_job.query_args = vars(
            ResubmitFormDataFactory(submit="download", names=self.case.get_members())
        )

    def _test_export_xlsx(self, database, mock_):
        self._set_mehari_mocker(mock_)

        self.export_job.query_args["database_select"] = database
        with file_export.CaseExporterXlsx(self.export_job, self.export_job.case) as exporter:
            result = exporter.generate()
        with tempfile.NamedTemporaryFile(suffix=".xlsx") as temp_file:
            temp_file.write(result)
            temp_file.flush()
            workbook = openpyxl.load_workbook(temp_file.name)
            # TODO: also test without commments
            self.assertEquals(workbook.sheetnames, ["Variants", "Comments", "Metadata"])
            variants_sheet = workbook["Variants"]
            arrs = [[cell.value for cell in row] for row in variants_sheet.rows]
            self._test_tabular(arrs, False, settings.VARFISH_BACKEND_URL_MEHARI, database)

    def _test_export_tsv(self, database, mock_):
        self._set_mehari_mocker(mock_)

        self.export_job.query_args["database_select"] = database
        with file_export.CaseExporterTsv(self.export_job, self.export_job.case) as exporter:
            result = str(exporter.generate(), "utf-8")
        arrs = [line.split("\t") for line in result.split("\n")]
        self._test_tabular(arrs, True, settings.VARFISH_BACKEND_URL_MEHARI, database)

    @patch("django.conf.settings.VARFISH_BACKEND_URL_MEHARI", "https://mehari.com")
    @Mocker()
    def test_export_tsv(self, mock_):
        self._test_export_tsv("refseq", mock_)

    @patch("django.conf.settings.VARFISH_BACKEND_URL_MEHARI", "https://mehari.com")
    @Mocker()
    def test_export_tsv_refseq(self, mock_):
        self._test_export_tsv("refseq", mock_)

    # TODO mehari does not provide ensembl transcripts ATM, only refseq
    # @patch("django.conf.settings.VARFISH_BACKEND_URL_MEHARI", "https://mehari.com")
    # @Mocker()
    # def test_export_tsv_ensembl(self, mock_):
    #     self._test_export_tsv("ensembl", mock_)

    def _test_tabular(self, arrs, has_trailing, mehari_enable, database):
        self.assertEquals(len(arrs), 4 + int(has_trailing))
        # TODO: also test without flags and comments
        if not mehari_enable:
            self.assertEquals(len(arrs[0]), 58)
        else:
            self.assertEquals(len(arrs[0]), 59)
        self.assertSequenceEqual(arrs[0][:3], ["Chromosome", "Position", "Reference bases"])
        self.assertSequenceEqual(
            arrs[0][-5:],
            [
                "%s Genotype" % self.case.pedigree[0]["patient"],
                "%s Gt. Quality" % self.case.pedigree[0]["patient"],
                "%s Alternative depth" % self.case.pedigree[0]["patient"],
                "%s Total depth" % self.case.pedigree[0]["patient"],
                "%s Alternate allele fraction" % self.case.pedigree[0]["patient"],
            ],
        )
        for i, small_var in enumerate(self.small_vars):
            self.assertSequenceEqual(
                arrs[i + 1][:3],
                ["chr" + small_var.chromosome, str(small_var.start), small_var.reference],
            )
            self.assertSequenceEqual(
                arrs[i + 1][-5:],
                list(
                    map(
                        str,
                        [
                            small_var.genotype[self.case.pedigree[-1]["patient"]]["gt"],
                            small_var.genotype[self.case.pedigree[-1]["patient"]]["gq"],
                            small_var.genotype[self.case.pedigree[-1]["patient"]]["ad"],
                            small_var.genotype[self.case.pedigree[-1]["patient"]]["dp"],
                            small_var.genotype[self.case.pedigree[-1]["patient"]]["ad"]
                            / small_var.genotype[self.case.pedigree[-1]["patient"]]["dp"],
                        ],
                    )
                ),
            )
            if i == 0:
                self.assertSequenceEqual(arrs[i + 1][-6:-5], ["9.89"])
            elif i == 2:
                self.assertSequenceEqual(arrs[i + 1][-6:-5], ["10.78"])
            self.assertSequenceEqual(
                [
                    arrs[i + 1][31][0:6],
                    arrs[i + 1][32][0:6],
                    arrs[i + 1][33][0:6],
                    arrs[i + 1][34][0:6],
                    arrs[i + 1][35][0:6],
                ],
                [
                    str(1 / 2 ** (i % 12) + 1.234)[0:6],
                    str(1 / 2 ** (i % 12))[0:6],
                    str(1 / 2 ** (i % 12))[0:6],
                    str(1 / 3 ** (i % 12))[0:6],
                    str((1 / 0.75) ** (i % 12))[0:6],
                ],
            )
            if mehari_enable:
                if database == "refseq":
                    self.assertEquals(
                        arrs[i + 1][38].replace("\n", "|"),
                        (
                            "NM_058167.2;three_prime_utr_exon_variant;p.(=);c.*60G>A|NM_194315.1;"
                            "three_prime_utr_exon_variant;p.(=);c.*60G>A"
                        ),
                    )
                else:
                    self.assertEquals(
                        arrs[i + 1][38].replace("\n", "|"),
                        (
                            "ENST_058167.2;three_prime_utr_exon_variant;p.(=);c.*60G>A|ENST_194315.1;"
                            "three_prime_utr_exon_variant;p.(=);c.*60G>A"
                        ),
                    )

            self.assertEquals(arrs[i + 1][36], "pathogenic")
            self.assertEquals(arrs[i + 1][37], "criteria provided, single committer")
        if has_trailing:
            self.assertSequenceEqual(arrs[4], [""])

    @patch("django.conf.settings.VARFISH_BACKEND_URL_MEHARI", "https://mehari.com")
    @Mocker()
    def test_export_vcf(self, mock_):
        self._set_mehari_mocker(mock_)
        with file_export.CaseExporterVcf(self.export_job, self.export_job.case) as exporter:
            result = exporter.generate()
        unzipped = gzip.GzipFile(fileobj=io.BytesIO(result), mode="rb").read()
        lines = str(unzipped, "utf-8").split("\n")
        header = [l for l in lines if l.startswith("#")]
        content = [l for l in lines if not l.startswith("#")]
        self.assertEquals(len(header), 31)
        self.assertEquals(header[0], "##fileformat=VCFv4.2")
        self.assertEquals(
            header[-1],
            "#CHROM\tPOS\tID\tREF\tALT\tQUAL\tFILTER\tINFO\tFORMAT\t%s"
            % self.case.pedigree[0]["patient"],
        )
        self.assertEquals(len(content), 4)
        for i, small_var in enumerate(self.small_vars):
            genotype = small_var.genotype[self.case.pedigree[0]["patient"]]
            self.assertEquals(
                content[i].split("\t"),
                list(
                    map(
                        str,
                        [
                            small_var.chromosome,
                            small_var.start,
                            ".",
                            small_var.reference,
                            small_var.alternative,
                            ".",
                            ".",
                            ".",
                            "GT:GQ:AD:DP",
                            "%s:%s:%s:%s"
                            % (genotype["gt"], genotype["gq"], genotype["ad"], genotype["dp"]),
                        ],
                    )
                ),
            )
        self.assertEquals(content[3], "")

    @patch("django.conf.settings.VARFISH_BACKEND_URL_MEHARI", "https://mehari.com")
    @Mocker()
    def test_export_xlsx(self, mock):
        self._test_export_xlsx("refseq", mock)

    @patch("django.conf.settings.VARFISH_BACKEND_URL_MEHARI", "https://mehari.com")
    @Mocker()
    def test_export_xlsx_refseq(self, mock):
        self._test_export_xlsx("refseq", mock)

    # TODO mehari does not provide ensembl transcripts ATM, only refseq
    # @patch("django.conf.settings.VARFISH_BACKEND_URL_MEHARI", "https://mehari.com")
    # @Mocker()
    # def test_export_xlsx_ensembl(self, mock):
    #     self._test_export_xlsx("ensembl", mock)


class ProjectExportTest(TestCase):
    """Base class for testing exports.

    Sets up the database fixtures for project, case, and small variants.
    """

    def setUp(self):
        self.superuser = self.make_user("superuser")
        self.case1, self.variant_set1, _ = CaseWithVariantSetFactory.get("small")
        self.project = CaseAwareProject.objects.get(pk=Project.objects.first().pk)
        self.case2, self.variant_set2, _ = CaseWithVariantSetFactory.get(
            "small", project=self.project
        )
        self.small_vars1 = [
            SmallVariantFactory(chromosome="1", variant_set=self.variant_set1),
            SmallVariantFactory(chromosome="2", variant_set=self.variant_set1),
            SmallVariantFactory(chromosome="3", variant_set=self.variant_set1),
        ]
        self.small_vars2 = [
            SmallVariantFactory(
                chromosome=self.small_vars1[0].chromosome,
                start=self.small_vars1[0].start,
                reference=self.small_vars1[0].reference,
                alternative=self.small_vars1[0].alternative,
                variant_set=self.variant_set2,
            ),
        ]
        self.bg_job = BackgroundJob.objects.create(
            name="job name",
            project=Project.objects.first(),
            job_type="variants.export_file_bg_job",
            user=self.superuser,
        )
        self.export_job = ExportProjectCasesFileBgJob.objects.create(
            project=self.bg_job.project,
            bg_job=self.bg_job,
            query_args=vars(
                ResubmitFormDataFactory(submit="download", names=self.project.get_members())
            ),
            file_type="xlsx",
        )

    @patch("django.conf.settings.VARFISH_BACKEND_URL_MEHARI", None)
    def test_export_tsv(self):
        with file_export.CaseExporterTsv(self.export_job, self.project) as exporter:
            result = str(exporter.generate(), "utf-8")
        arrs = [line.split("\t") for line in result.split("\n")]
        self._test_tabular(arrs, True)

    def _test_tabular(self, arrs, has_trailing):
        self.assertEquals(len(arrs), 5 + int(has_trailing))
        # TODO: also test without flags and comments
        if settings.VARFISH_BACKEND_URL_MEHARI:
            self.assertEquals(len(arrs[0]), 59)
        else:
            self.assertEquals(len(arrs[0]), 58)
        self.assertSequenceEqual(arrs[0][:3], ["Sample", "Chromosome", "Position"])
        self.assertEqual(arrs[0][-1], "sample Alternate allele fraction")
        members = sorted(self.project.get_members())
        for i, small_var in enumerate(
            sorted(
                self.small_vars1 + self.small_vars2,
                key=lambda x: (x.chromosome_no, x.start, members.index(list(x.genotype.keys())[0])),
            )
        ):
            member = list(small_var.genotype.keys())[0]
            self.assertSequenceEqual(
                arrs[i + 1][:3],
                [
                    member,
                    "chr" + small_var.chromosome,
                    str(small_var.start),
                ],
            )
            self.assertSequenceEqual(
                arrs[i + 1][-5:],
                list(
                    map(
                        str,
                        [
                            small_var.genotype[member]["gt"],
                            small_var.genotype[member]["gq"],
                            small_var.genotype[member]["ad"],
                            small_var.genotype[member]["dp"],
                            small_var.genotype[member]["ad"] / small_var.genotype[member]["dp"],
                        ],
                    )
                ),
            )
        if has_trailing:
            self.assertSequenceEqual(arrs[5], [""])

    @patch("django.conf.settings.VARFISH_BACKEND_URL_MEHARI", None)
    def test_export_vcf(self):
        with file_export.CaseExporterVcf(self.export_job, self.project) as exporter:
            result = exporter.generate()
        unzipped = gzip.GzipFile(fileobj=io.BytesIO(result), mode="rb").read()
        lines = str(unzipped, "utf-8").split("\n")
        header = [l for l in lines if l.startswith("#")]
        content = [l for l in lines if not l.startswith("#")]
        member1 = self.case1.pedigree[0]["patient"]
        member2 = self.case2.pedigree[0]["patient"]
        self.assertEquals(len(header), 31)
        self.assertEquals(header[0], "##fileformat=VCFv4.2")
        self.assertEquals(
            header[-1],
            "#CHROM\tPOS\tID\tREF\tALT\tQUAL\tFILTER\tINFO\tFORMAT\t%s\t%s" % (member1, member2),
        )
        self.assertEquals(len(content), 4)
        self.assertEquals(
            content[0].split("\t"),
            [
                self.small_vars1[0].chromosome,
                str(self.small_vars1[0].start),
                ".",
                self.small_vars1[0].reference,
                self.small_vars1[0].alternative,
                ".",
                ".",
                ".",
                "GT:GQ:AD:DP",
                "%s:%s:%s:%s"
                % (
                    self.small_vars1[0].genotype[member1]["gt"],
                    self.small_vars1[0].genotype[member1]["gq"],
                    self.small_vars1[0].genotype[member1]["ad"],
                    self.small_vars1[0].genotype[member1]["dp"],
                ),
                "%s:%s:%s:%s"
                % (
                    self.small_vars2[0].genotype[member2]["gt"],
                    self.small_vars2[0].genotype[member2]["gq"],
                    self.small_vars2[0].genotype[member2]["ad"],
                    self.small_vars2[0].genotype[member2]["dp"],
                ),
            ],
        )
        self.assertEquals(
            content[1].split("\t"),
            [
                self.small_vars1[1].chromosome,
                str(self.small_vars1[1].start),
                ".",
                self.small_vars1[1].reference,
                self.small_vars1[1].alternative,
                ".",
                ".",
                ".",
                "GT:GQ:AD:DP",
                "%s:%s:%s:%s"
                % (
                    self.small_vars1[1].genotype[member1]["gt"],
                    self.small_vars1[1].genotype[member1]["gq"],
                    self.small_vars1[1].genotype[member1]["ad"],
                    self.small_vars1[1].genotype[member1]["dp"],
                ),
                "./.:.:.:.",
            ],
        )
        self.assertEquals(
            content[2].split("\t"),
            [
                self.small_vars1[2].chromosome,
                str(self.small_vars1[2].start),
                ".",
                self.small_vars1[2].reference,
                self.small_vars1[2].alternative,
                ".",
                ".",
                ".",
                "GT:GQ:AD:DP",
                "%s:%s:%s:%s"
                % (
                    self.small_vars1[2].genotype[member1]["gt"],
                    self.small_vars1[2].genotype[member1]["gq"],
                    self.small_vars1[2].genotype[member1]["ad"],
                    self.small_vars1[2].genotype[member1]["dp"],
                ),
                "./.:.:.:.",
            ],
        )
        self.assertEquals(content[3], "")

    @patch("django.conf.settings.VARFISH_BACKEND_URL_MEHARI", None)
    def test_export_xlsx(self):
        with file_export.CaseExporterXlsx(self.export_job, self.project) as exporter:
            result = exporter.generate()
        with tempfile.NamedTemporaryFile(suffix=".xlsx") as temp_file:
            temp_file.write(result)
            temp_file.flush()
            workbook = openpyxl.load_workbook(temp_file.name)
            # TODO: also test without commments
            self.assertEquals(workbook.sheetnames, ["Variants", "Comments", "Metadata"])
            variants_sheet = workbook["Variants"]
            arrs = [[cell.value for cell in row] for row in variants_sheet.rows]
            self._test_tabular(arrs, False)


class CohortExporterTest(TestCohortBase):
    def _create_bgjob(self, user, cohort):
        return ExportProjectCasesFileBgJobFactory(
            user=user,
            project=cohort.project,
            query_args=vars(
                ResubmitFormDataFactory(submit="download", names=cohort.get_members(user))
            ),
            file_type="xlsx",
        )

    @patch("django.conf.settings.VARFISH_BACKEND_URL_MEHARI", None)
    def test_export_tsv_as_superuser(self):
        user = self.superuser
        project = self.project1
        cohort = self._create_cohort_all_possible_cases(user, project)
        bgjob = self._create_bgjob(user, cohort)
        with file_export.CaseExporterTsv(bgjob, cohort) as exporter:
            result = str(exporter.generate(), "utf-8")
        arrs = [line.split("\t") for line in result.split("\n")]
        self._test_tabular(
            arrs,
            16,
            True,
            self.project1_case1_smallvars
            + self.project1_case2_smallvars
            + self.project2_case1_smallvars
            + self.project2_case2_smallvars,
        )

    @patch("django.conf.settings.VARFISH_BACKEND_URL_MEHARI", None)
    def test_export_tsv_as_contributor(self):
        user = self.contributor
        project = self.project2
        cohort = self._create_cohort_all_possible_cases(user, project)
        bgjob = self._create_bgjob(user, cohort)
        with file_export.CaseExporterTsv(bgjob, cohort) as exporter:
            result = str(exporter.generate(), "utf-8")
        arrs = [line.split("\t") for line in result.split("\n")]
        self._test_tabular(
            arrs,
            13,
            True,
            self.project2_case1_smallvars + self.project2_case2_smallvars,
        )

    @patch("django.conf.settings.VARFISH_BACKEND_URL_MEHARI", None)
    def test_export_tsv_as_superuser_for_cohort_by_contributor(self):
        user = self.superuser
        project = self.project2
        cohort = self._create_cohort_all_possible_cases(self.contributor, project)
        bgjob = self._create_bgjob(user, cohort)
        with file_export.CaseExporterTsv(bgjob, cohort) as exporter:
            result = str(exporter.generate(), "utf-8")
        arrs = [line.split("\t") for line in result.split("\n")]
        self._test_tabular(
            arrs,
            13,
            True,
            self.project2_case1_smallvars + self.project2_case2_smallvars,
        )

    @patch("django.conf.settings.VARFISH_BACKEND_URL_MEHARI", None)
    def test_export_tsv_as_contributor_for_cohort_by_superuser(self):
        user = self.contributor
        project = self.project2
        cohort = self._create_cohort_all_possible_cases(self.superuser, project)
        bgjob = self._create_bgjob(user, cohort)
        with file_export.CaseExporterTsv(bgjob, cohort) as exporter:
            result = str(exporter.generate(), "utf-8")
        arrs = [line.split("\t") for line in result.split("\n")]
        self._test_tabular(
            arrs,
            13,
            True,
            self.project2_case1_smallvars + self.project2_case2_smallvars,
        )

    @patch("django.conf.settings.VARFISH_BACKEND_URL_MEHARI", None)
    def _test_tabular(self, arrs, ref, has_trailing, smallvars):
        self.assertEquals(len(arrs), ref + int(has_trailing))
        # TODO: also test without flags and comments
        if settings.VARFISH_BACKEND_URL_MEHARI:
            self.assertEquals(len(arrs[0]), 59)
        else:
            self.assertEquals(len(arrs[0]), 58)
        self.assertSequenceEqual(arrs[0][:3], ["Sample", "Chromosome", "Position"])
        self.assertEqual(arrs[0][-1], "sample Alternate allele fraction")
        for i, small_var in enumerate(sorted(smallvars, key=lambda x: (x.chromosome_no, x.start))):
            member = Case.objects.get(id=small_var.case_id).get_members()[0]
            self.assertSequenceEqual(
                arrs[i + 1][:3],
                [
                    member,
                    "chr" + small_var.chromosome,
                    str(small_var.start),
                ],
            )
            self.assertSequenceEqual(
                arrs[i + 1][-5:],
                list(
                    map(
                        str,
                        [
                            small_var.genotype[member]["gt"],
                            small_var.genotype[member]["gq"],
                            small_var.genotype[member]["ad"],
                            small_var.genotype[member]["dp"],
                            small_var.genotype[member]["ad"] / small_var.genotype[member]["dp"],
                        ],
                    )
                ),
            )
        if has_trailing:
            self.assertSequenceEqual(arrs[ref], [""])

    def _test_vcf(self, result, ref, smallvars, cohort, user):
        unzipped = gzip.GzipFile(fileobj=io.BytesIO(result), mode="rb").read()
        lines = str(unzipped, "utf-8").split("\n")
        header = [l for l in lines if l.startswith("#")]
        content = [l for l in lines if not l.startswith("#")]
        members = cohort.get_members(user)
        self.assertEquals(len(header), 31)
        self.assertEquals(header[0], "##fileformat=VCFv4.2")
        self.assertEquals(
            header[-1],
            "#CHROM\tPOS\tID\tREF\tALT\tQUAL\tFILTER\tINFO\tFORMAT\t%s" % "\t".join(members),
        )
        self.assertEquals(len(content), ref)
        vcf_vars = {}
        for i in smallvars:
            vcf_vars[
                (
                    i.chromosome_no,
                    i.chromosome,
                    str(i.start),
                    ".",
                    i.reference,
                    i.alternative,
                    ".",
                    ".",
                    ".",
                    "GT:GQ:AD:DP",
                )
            ] = [
                (
                    "%s:%s:%s:%s"
                    % (
                        i.genotype[m]["gt"],
                        i.genotype[m]["gq"],
                        i.genotype[m]["ad"],
                        i.genotype[m]["dp"],
                    )
                    if m in i.genotype
                    else "./.:.:.:."
                )
                for m in members
            ]
        for i, var in enumerate(sorted(vcf_vars, key=lambda x: (x[0], int(x[2])))):
            self.assertEqual(content[i].split("\t"), list(var[1:]) + vcf_vars[var])
        self.assertEquals(content[ref - 1], "")

    @patch("django.conf.settings.VARFISH_BACKEND_URL_MEHARI", None)
    def test_export_vcf_as_superuser(self):
        user = self.superuser
        project = self.project1
        cohort = self._create_cohort_all_possible_cases(user, project)
        bgjob = self._create_bgjob(user, cohort)
        with file_export.CaseExporterVcf(bgjob, cohort) as exporter:
            result = exporter.generate()
        self._test_vcf(
            result,
            16,
            self.project1_case1_smallvars
            + self.project1_case2_smallvars
            + self.project2_case1_smallvars
            + self.project2_case2_smallvars,
            cohort,
            user,
        )

    @patch("django.conf.settings.VARFISH_BACKEND_URL_MEHARI", None)
    def test_export_vcf_as_contributor(self):
        user = self.contributor
        project = self.project2
        cohort = self._create_cohort_all_possible_cases(user, project)
        bgjob = self._create_bgjob(user, cohort)
        with file_export.CaseExporterVcf(bgjob, cohort) as exporter:
            result = exporter.generate()
        self._test_vcf(
            result,
            13,
            self.project2_case1_smallvars + self.project2_case2_smallvars,
            cohort,
            user,
        )

    @patch("django.conf.settings.VARFISH_BACKEND_URL_MEHARI", None)
    def test_export_vcf_as_superuser_for_cohort_by_contributor(self):
        user = self.superuser
        project = self.project2
        cohort = self._create_cohort_all_possible_cases(self.contributor, project)
        bgjob = self._create_bgjob(user, cohort)
        with file_export.CaseExporterVcf(bgjob, cohort) as exporter:
            result = exporter.generate()
        self._test_vcf(
            result,
            13,
            self.project2_case1_smallvars + self.project2_case2_smallvars,
            cohort,
            user,
        )

    @patch("django.conf.settings.VARFISH_BACKEND_URL_MEHARI", None)
    def test_export_vcf_as_contributor_for_cohort_by_superuser(self):
        user = self.contributor
        project = self.project2
        cohort = self._create_cohort_all_possible_cases(self.superuser, project)
        bgjob = self._create_bgjob(user, cohort)
        with file_export.CaseExporterVcf(bgjob, cohort) as exporter:
            result = exporter.generate()
        self._test_vcf(
            result,
            13,
            self.project2_case1_smallvars + self.project2_case2_smallvars,
            cohort,
            user,
        )

    @patch("django.conf.settings.VARFISH_BACKEND_URL_MEHARI", None)
    def test_export_xlsx_as_superuser(self):
        user = self.superuser
        project = self.project1
        cohort = self._create_cohort_all_possible_cases(user, project)
        bgjob = self._create_bgjob(user, cohort)
        with file_export.CaseExporterXlsx(bgjob, cohort) as exporter:
            result = exporter.generate()
        with tempfile.NamedTemporaryFile(suffix=".xlsx") as temp_file:
            temp_file.write(result)
            temp_file.flush()
            workbook = openpyxl.load_workbook(temp_file.name)
            # TODO: also test without commments
            self.assertEquals(workbook.sheetnames, ["Variants", "Comments", "Metadata"])
            variants_sheet = workbook["Variants"]
            arrs = [[cell.value for cell in row] for row in variants_sheet.rows]
            self._test_tabular(
                arrs,
                16,
                False,
                self.project1_case1_smallvars
                + self.project1_case2_smallvars
                + self.project2_case1_smallvars
                + self.project2_case2_smallvars,
            )

    @patch("django.conf.settings.VARFISH_BACKEND_URL_MEHARI", None)
    def test_export_xlsx_as_contributor(self):
        user = self.contributor
        project = self.project2
        cohort = self._create_cohort_all_possible_cases(user, project)
        bgjob = self._create_bgjob(user, cohort)
        with file_export.CaseExporterXlsx(bgjob, cohort) as exporter:
            result = exporter.generate()
        with tempfile.NamedTemporaryFile(suffix=".xlsx") as temp_file:
            temp_file.write(result)
            temp_file.flush()
            workbook = openpyxl.load_workbook(temp_file.name)
            # TODO: also test without commments
            self.assertEquals(workbook.sheetnames, ["Variants", "Comments", "Metadata"])
            variants_sheet = workbook["Variants"]
            arrs = [[cell.value for cell in row] for row in variants_sheet.rows]
            self._test_tabular(
                arrs,
                13,
                False,
                self.project2_case1_smallvars + self.project2_case2_smallvars,
            )

    @patch("django.conf.settings.VARFISH_BACKEND_URL_MEHARI", None)
    def test_export_xlsx_as_superuser_for_cohort_by_contributor(self):
        user = self.superuser
        project = self.project2
        cohort = self._create_cohort_all_possible_cases(self.contributor, project)
        bgjob = self._create_bgjob(user, cohort)
        with file_export.CaseExporterXlsx(bgjob, cohort) as exporter:
            result = exporter.generate()
        with tempfile.NamedTemporaryFile(suffix=".xlsx") as temp_file:
            temp_file.write(result)
            temp_file.flush()
            workbook = openpyxl.load_workbook(temp_file.name)
            # TODO: also test without commments
            self.assertEquals(workbook.sheetnames, ["Variants", "Comments", "Metadata"])
            variants_sheet = workbook["Variants"]
            arrs = [[cell.value for cell in row] for row in variants_sheet.rows]
            self._test_tabular(
                arrs,
                13,
                False,
                self.project2_case1_smallvars + self.project2_case2_smallvars,
            )

    @patch("django.conf.settings.VARFISH_BACKEND_URL_MEHARI", None)
    def test_export_xlsx_as_contributor_for_cohort_by_superuser(self):
        user = self.contributor
        project = self.project2
        cohort = self._create_cohort_all_possible_cases(self.superuser, project)
        bgjob = self._create_bgjob(user, cohort)
        with file_export.CaseExporterXlsx(bgjob, cohort) as exporter:
            result = exporter.generate()
        with tempfile.NamedTemporaryFile(suffix=".xlsx") as temp_file:
            temp_file.write(result)
            temp_file.flush()
            workbook = openpyxl.load_workbook(temp_file.name)
            # TODO: also test without commments
            self.assertEquals(workbook.sheetnames, ["Variants", "Comments", "Metadata"])
            variants_sheet = workbook["Variants"]
            arrs = [[cell.value for cell in row] for row in variants_sheet.rows]
            self._test_tabular(
                arrs,
                13,
                False,
                self.project2_case1_smallvars + self.project2_case2_smallvars,
            )


def _fake_generate(_self):
    """Helper used for patching away ``CaseExporter*.generate``."""
    return bytes("test bytes", "utf-8")


class ExportCaseTest(ExportTestBase):
    """Test the ``export_case()`` function.

    We mock out the ``CaseExporter*`` class in the spirit of testing just the export driver code unit in
    ``export_case()``.
    """

    def _run_test(self, file_type):
        # Set the file type that we want to test into the export file job
        self.export_job.file_type = file_type
        self.export_job.save()
        # Run code under test
        file_export.export_case(self.export_job)
        # Check immediate result
        self.assertIsNotNone(self.export_job.export_result)
        self.assertEquals(self.export_job.export_result.payload, _fake_generate(self))
        # Check side effects
        self.assertEquals(ProjectEvent.objects.count(), 1)

    @patch.object(file_export.CaseExporterTsv, "generate", new=_fake_generate, create=True)
    def test_export_tsv(self):
        self._run_test("tsv")

    @patch.object(file_export.CaseExporterXlsx, "generate", new=_fake_generate, create=True)
    def test_export_xlsx(self):
        self._run_test("xlsx")

    @patch.object(file_export.CaseExporterVcf, "generate", new=_fake_generate, create=True)
    def test_export_vcf(self):
        self._run_test("vcf")


class ClearExpiredExportedFilesTest(ExportTestBase):
    """Test the ``clear_expired_exported_files()`` function."""

    def testWithExpired(self):
        file_export.ExportFileJobResult.objects.create(
            job=self.export_job, expiry_time=timezone.now() - timedelta(days=1)
        )
        self.assertEquals(file_export.ExportFileJobResult.objects.count(), 1)
        file_export.clear_expired_exported_files()
        self.assertEquals(file_export.ExportFileJobResult.objects.count(), 0)

    def testWithNonExpired(self):
        file_export.ExportFileJobResult.objects.create(
            job=self.export_job, expiry_time=timezone.now() + timedelta(days=1)
        )
        self.assertEquals(file_export.ExportFileJobResult.objects.count(), 1)
        file_export.clear_expired_exported_files()
        self.assertEquals(file_export.ExportFileJobResult.objects.count(), 1)
